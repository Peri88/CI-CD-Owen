#!/usr/bin/env python3
import argparse
from datetime import datetime, date
from pathlib import Path
import re
import os
import zipfile
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell


def read_excel_with_retry(path, sheet_name, header=None, engine="openpyxl", retries=5, delay=1.0):
    """Read Excel with a few retries to handle partial uploads."""
    last_err = None
    for _ in range(retries):
        try:
            return pd.read_excel(path, sheet_name=sheet_name, header=header, engine=engine)
        except Exception as e:
            last_err = e
            import time
            time.sleep(delay)
    raise last_err

HEADER_ROW = [
    "State", "Policy", "Job", "Schedule", "Client", "Media", "Server",
    "Start", "Time", "Elapsed", "Time.1", "End", "Time.2", "Unit"
]

# Column indices in Export1.xlsx (0-based)
COL_POLICY = 4
COL_START_Y = 8
COL_START_M = 9
COL_START_D = 10
COL_START_AMPM = 11
COL_START_TIME = 12
COL_ELAPSED = 13
COL_END_Y = 14
COL_END_M = 15
COL_END_D = 16
COL_END_AMPM = 17
COL_END_TIME = 18
COL_STORAGE_UNIT = 19
COL_UNIT = 21

REPORT_GLOB = "/home/owen/벽산 리포트_백업상태_최종(양식)_*.xlsx"


TEMPLATE_PATH = "/home/owen/벽산 리포트_백업상태_최종(양식).xlsx"


def build_parsed_df(export1_path: str, include_all_dates: bool = False) -> pd.DataFrame:
    raw = read_excel_with_retry(export1_path, sheet_name="Export1", header=None)
    raw = raw.iloc[1:].reset_index(drop=True)

    rows = []
    for _, r in raw.iterrows():
        policy = r[COL_POLICY]
        if pd.isna(policy):
            continue
        row = {
            "Policy": str(policy).strip(),
            "Start_Y": r[COL_START_Y],
            "Start_M": r[COL_START_M],
            "Start_D": r[COL_START_D],
            "Start_AMPM": r[COL_START_AMPM],
            "Start_Time": r[COL_START_TIME],
            "Elapsed": r[COL_ELAPSED],
            "End_Y": r[COL_END_Y],
            "End_M": r[COL_END_M],
            "End_D": r[COL_END_D],
            "End_AMPM": r[COL_END_AMPM],
            "End_Time": r[COL_END_TIME],
            "Storage_Unit": r[COL_STORAGE_UNIT],
            "Unit": r[COL_UNIT],
        }
        rows.append(row)

    if include_all_dates:
        filtered = rows
    else:
        # latest date per policy
        latest_date = {}
        for row in rows:
            try:
                d = datetime(int(row["End_Y"]), int(row["End_M"]), int(row["End_D"])).date()
            except Exception:
                continue
            pol = row["Policy"]
            if pol not in latest_date or d > latest_date[pol]:
                latest_date[pol] = d

        # filter to latest date rows only
        filtered = []
        for row in rows:
            pol = row["Policy"]
            try:
                d = datetime(int(row["End_Y"]), int(row["End_M"]), int(row["End_D"])).date()
            except Exception:
                continue
            if latest_date.get(pol) == d:
                filtered.append(row)

    # HZDB_MSSQL split labeling
    for row in filtered:
        if row["Policy"] == "HZDB_MSSQL":
            try:
                unit = int(str(row["Unit"]).replace(",", ""))
            except Exception:
                unit = None
            if unit is None:
                continue
            if 8000 <= unit < 10000:
                row["Policy"] = "HZDB_MSSQL_ReportServer"
            elif 1000000 <= unit < 2000000:
                row["Policy"] = "HZDB_MSSQL_SMS"
            else:
                row["Policy"] = "HZDB_MSSQL_NEOE"

    # build output rows with header row exactly like test.xlsx
    out_rows = [HEADER_ROW]
    for row in filtered:
        out_rows.append([
            row["Policy"],
            row["Start_Y"],
            row["Start_M"],
            row["Start_D"],
            row["Start_AMPM"],
            row["Start_Time"],
            row["Elapsed"],
            row["End_Y"],
            row["End_M"],
            row["End_D"],
            row["End_AMPM"],
            row["End_Time"],
            row["Storage_Unit"],
            row["Unit"],
        ])

    return pd.DataFrame(out_rows)


def _parse_unit_from_cell(cell_value):
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float)):
        return float(cell_value)
    if isinstance(cell_value, str):
        m = re.match(r"=\s*(\d+)\s*/\s*\(1024\*1024\)", cell_value)
        if m:
            try:
                unit = int(m.group(1))
                return unit / 1024 / 1024
            except Exception:
                return None
        # try numeric string
        try:
            return float(cell_value)
        except Exception:
            return None
    return None


def _format_gb(val: float) -> str:
    # keep two decimals if needed, else integer
    if val is None:
        return ""
    if abs(val - round(val)) < 0.005:
        return str(int(round(val)))
    return f"{val:.2f}"


def _find_previous_report(current_report: str) -> str | None:
    # find latest dated report before today (by filename tag)
    candidates = []
    for p in Path("/home/owen").glob("벽산 리포트_백업상태_최종(양식)_*.xlsx"):
        name = p.name
        m = re.search(r"_(\d{8})(?:_\d{6})?\.xlsx$", name)
        if not m:
            continue
        tag = m.group(1)
        try:
            d = datetime.strptime(tag, "%Y%m%d").date()
        except Exception:
            continue
        candidates.append((d, str(p)))

    if not candidates:
        return None

    candidates.sort()
    # choose latest before today, else latest overall
    today = date.today()
    before = [c for c in candidates if c[0] < today]
    if before:
        return before[-1][1]
    return candidates[-1][1]


def _read_previous_values(prev_report: str) -> dict:
    wb = openpyxl.load_workbook(prev_report, data_only=False)
    ws = wb["백업상태 점검_일일점검"]

    prev = {}
    # policy rows (col C)
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=3, max_col=3):
        cell = row[0]
        v = cell.value
        if isinstance(v, str):
            pol = v.strip()
            gb = _parse_unit_from_cell(ws.cell(cell.row, 5).value)
            if gb is not None:
                prev[pol] = gb

    # HZDB split rows by label in col D
    label_map = {
        "ReportServer": "HZDB_MSSQL_ReportServer",
        "SMS": "HZDB_MSSQL_SMS",
        "NEOE": "HZDB_MSSQL_NEOE",
    }
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=4, max_col=4):
        cell = row[0]
        v = cell.value
        if isinstance(v, str) and v.strip() in label_map:
            key = label_map[v.strip()]
            gb = _parse_unit_from_cell(ws.cell(cell.row, 5).value)
            if gb is not None:
                prev[key] = gb

    return prev


def restore_sheet1_assets(template_path: str, report_path: str):
    # Preserve Sheet1 drawings/images by copying parts from the template.
    if not Path(template_path).exists():
        print(f"[WARN] template missing: {template_path}")
        return

    def should_override(name: str) -> bool:
        if name in ("[Content_Types].xml", "xl/worksheets/sheet1.xml", "xl/worksheets/_rels/sheet1.xml.rels"):
            return True
        if name.startswith(("xl/media/", "xl/drawings/", "xl/ink/", "xl/printerSettings/")):
            return True
        return False

    def should_copy_missing(name: str) -> bool:
        if should_override(name):
            return True
        if name in ("xl/sharedStrings.xml", "xl/calcChain.xml", "xl/worksheets/_rels/sheet2.xml.rels"):
            return True
        return False

    tmp_path = report_path + ".tmp"
    with zipfile.ZipFile(report_path, "r") as z_out, zipfile.ZipFile(template_path, "r") as z_tpl:
        out_names = set(z_out.namelist())
        tpl_names = set(z_tpl.namelist())
        with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as z_new:
            for name in out_names:
                if should_override(name) and name in tpl_names:
                    z_new.writestr(name, z_tpl.read(name))
                else:
                    z_new.writestr(name, z_out.read(name))
            for name in tpl_names - out_names:
                if should_copy_missing(name):
                    z_new.writestr(name, z_tpl.read(name))
    os.replace(tmp_path, report_path)


def update_report(report_path: str, parsed_path: str):
    parsed = pd.read_excel(parsed_path, sheet_name="Export1", header=None)
    parsed = parsed.iloc[1:].reset_index(drop=True)

    policy_col = 0
    unit_col = 13
    parsed[unit_col] = pd.to_numeric(parsed[unit_col], errors="coerce").fillna(0)

    agg = parsed.groupby(policy_col)[unit_col].sum()

    # current gb values by key
    current_gb = {k: v / 1024 / 1024 for k, v in agg.items()}

    prev_report = _find_previous_report(report_path)
    prev_values = _read_previous_values(prev_report) if prev_report else {}

    wb = openpyxl.load_workbook(report_path)
    ws = wb["백업상태 점검_일일점검"]

    # update inspection date
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "점검일시" in v:
                cell.value = f"점검일시 : {date.today().isoformat()}"

    # fill backup volume for policy rows (col C => col E)
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=3, max_col=3):
        cell = row[0]
        v = cell.value
        if isinstance(v, str):
            pol = v.strip()
            if pol in agg:
                unit_sum = int(agg[pol])
                ws.cell(cell.row, 5).value = f"={unit_sum}/(1024*1024)"

    # HZDB_MSSQL split rows in col D
    label_map = {
        "ReportServer": "HZDB_MSSQL_ReportServer",
        "SMS": "HZDB_MSSQL_SMS",
        "NEOE": "HZDB_MSSQL_NEOE",
    }
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=4, max_col=4):
        cell = row[0]
        v = cell.value
        if isinstance(v, str) and v.strip() in label_map:
            pol = label_map[v.strip()]
            if pol in agg:
                unit_sum = int(agg[pol])
                ws.cell(cell.row, 5).value = f"={unit_sum}/(1024*1024)"

    # update remarks if delta >= 10GB, except ERP-APP
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=3, max_col=8):
        policy_cell = row[0]  # col C
        remark_cell = row[5]  # col H
        pol = policy_cell.value.strip() if isinstance(policy_cell.value, str) else ""
        if not pol:
            continue
        if isinstance(remark_cell, MergedCell):
            continue
        if pol == "ERP-APP":
            # keep existing remark
            continue
        if pol in current_gb and pol in prev_values:
            cur = current_gb[pol]
            prev = prev_values[pol]
            diff = cur - prev
            if abs(diff) >= 10:
                trend = "증가" if diff > 0 else "감소"
                remark_cell.value = f"{_format_gb(prev)}GB -> {_format_gb(cur)}GB ({_format_gb(abs(diff))}GB{trend})"

    # HZDB split remark rows by label in col D
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=4, max_col=8):
        label_cell = row[0]  # col D
        remark_cell = row[4]  # col H
        label = label_cell.value.strip() if isinstance(label_cell.value, str) else ""
        if label in label_map:
            if isinstance(remark_cell, MergedCell):
                continue
            key = label_map[label]
            if key in current_gb and key in prev_values:
                cur = current_gb[key]
                prev = prev_values[key]
                diff = cur - prev
                if abs(diff) >= 10:
                    trend = "증가" if diff > 0 else "감소"
                    remark_cell.value = f"{_format_gb(prev)}GB -> {_format_gb(cur)}GB ({_format_gb(abs(diff))}GB{trend})"

    wb.save(report_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--export1", default="/home/owen/Export1.xlsx")
    ap.add_argument("--parsed", required=True)
    ap.add_argument("--report", required=True)
    ap.add_argument("--all-dates", action="store_true", help="Include all dates (no latest-date filtering)")
    args = ap.parse_args()

    parsed_df = build_parsed_df(args.export1, include_all_dates=args.all_dates)
    with pd.ExcelWriter(args.parsed, engine="openpyxl") as writer:
        parsed_df.to_excel(writer, sheet_name="Export1", header=False, index=False)

    update_report(args.report, args.parsed)
    restore_sheet1_assets(TEMPLATE_PATH, args.report)

    print(f"[OK] parsed: {args.parsed}")
    print(f"[OK] report updated: {args.report}")


if __name__ == "__main__":
    main()
